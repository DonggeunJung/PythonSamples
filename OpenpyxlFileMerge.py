from copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook

def copy_ws(ws_src, wb_dst, new_sheet_name) :
    wb_dst.create_sheet(new_sheet_name)
    new_sheet = wb_dst[new_sheet_name]
    for row in ws_src :
        for cell in row :
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

    if ws_src.merged_cells :
        new_sheet.merge_cells(str(ws_src.merged_cells))

wb_dst = Workbook()
src_file_list = ['./docs/sheet_copy1.xlsx', './docs/sheet_copy2.xlsx']

for src_file in src_file_list :
    wb_src = load_workbook(src_file)
    for ws_src in wb_src.worksheets :
        new_sheet_name = src_file[7:].split('.xlsx')[0] + '_' + ws_src.title
        copy_ws(ws_src, wb_dst, new_sheet_name)

wb_dst.remove(wb_dst['Sheet'])
wb_dst.save('merge.xlsx')

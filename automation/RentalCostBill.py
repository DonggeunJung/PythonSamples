import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Side, Border, numbers

invoice_table_header_list = [
    ['A', 'Consumer', 20], ['B', 'Service date', 12], ['C', 'Service type', 40],
    ['D', 'Manager', 12], ['E', 'Fee/Free', 10], ['F', 'Cost', 15],
    ['G', 'Tax', 15], ['H', 'Total', 15]
]

def makeInvoiceWorkbook() :
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = 'Service invoice'
    ws['A2'] = 'Rental copy machine service invoice'
    ws['A4'] = 'Service period : 2021.10'

    for header in invoice_table_header_list :
        ws[header[0] + '5'] = header[1]
    ws['A8'] = 'Above cost is requested.'
    ws['A9'] = '2021.10.29'
    ws['A10'] = 'Total rental Copyman'
    return wb

def stylizeInvoiceForm(ws, item_start_row_no, item_row_no) :
    side = Side(color='CCCCCC', border_style='medium')
    border_style = Border(left = side, right = side, top=side, bottom=side)

    for header in invoice_table_header_list :
        ws.column_dimensions[header[0]].width = header[2]
        ws[header[0] + '5'].font = Font(sz='12', bold=True)
        ws[header[0] + '5'].fill = PatternFill(patternType='solid',
                                               start_color='ffd663')
        ws[header[0] + '5'].border = border_style
        ws[header[0] + '5'].alignment\
            = Alignment(horizontal='center', vertical='center')

        for row_no in range(item_start_row_no, item_row_no + 1) :
            ws[f'{header[0]}{row_no}'].font = Font(sz='11', bold=False)
            ws[f'{header[0]}{row_no}'].border = border_style
            if header[0] in ['A', 'B', 'C', 'D', 'E'] :
                ws[f'{header[0]}{row_no}'].alignment \
                    = Alignment(horizontal='center', vertical='center')
            else :
                ws[f'{header[0]}{row_no}'].alignment \
                    = Alignment(horizontal='right', vertical='center')
                ws[f'{header[0]}{row_no}'].number_format\
                    = numbers.BUILTIN_FORMATS[37]

        ws[f'{header[0]}{item_row_no}'].font = Font(bold=True)
        ws[f'{header[0]}{item_row_no}'].fill \
            = PatternFill(patternType='solid', start_color='eeeeee')

    ws.merge_cells('A2:H2')
    ws.merge_cells('A4:H4')
    ws.merge_cells(f'B{item_row_no}:E{item_row_no}')
    ws.merge_cells(f'A{item_row_no+2}:H{item_row_no+2}')
    ws.merge_cells(f'A{item_row_no+3}:H{item_row_no+3}')
    ws.merge_cells(f'A{item_row_no+4}:H{item_row_no+4}')

    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].font = Font(sz='20', bold=True)
    ws['A4'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{item_row_no + 2}'].alignment \
        = Alignment(horizontal='center', vertical='center')
    ws[f'A{item_row_no + 2}'].font = Font(sz='16', bold=True)
    ws[f'A{item_row_no + 3}'].alignment \
        = Alignment(horizontal='center', vertical='center')
    ws[f'A{item_row_no + 3}'].font = Font(sz='12', bold=True)
    ws[f'A{item_row_no + 4}'].alignment \
        = Alignment(horizontal='right', vertical='center')
    ws[f'A{item_row_no + 4}'].font = Font(sz='14', bold=True)

    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[item_row_no + 2].height = 40
    ws.row_dimensions[item_row_no + 3].height = 20
    ws.row_dimensions[item_row_no + 4].height = 40

    for row_no in range(item_start_row_no, item_row_no+1) :
        ws.row_dimensions[row_no].height = 25

df_src = pd.read_excel('./docs/service_list.xlsx', sheet_name=None, engine='openpyxl')
df = pd.concat(df_src, ignore_index=True)
# print(df)
# df_filtered = df.loc[df['고객사명'] == '타오롱글로벌']
# print(df_filtered)

customer_list = df['고객사명'].unique()
# print(f'Count of customers : {len(customer_list)}')
# print(f'Customers list : {customer_list}')
# wb = makeInvoiceWorkbook()
# wb.save('service_invoice.xlsx')
for customer in customer_list :
    wb = makeInvoiceWorkbook()
    ws = wb['Service invoice']
    this_customer_items = df.loc[df['고객사명'] == customer]
    this_customer_items = this_customer_items.sort_values(by=['방문일자', '기사'])
    #print(this_customer_items)
    #break
    item_row_no = 6
    for idx in range(len(this_customer_items)) :
        ws.insert_rows(item_row_no)
        ws[f'A{item_row_no}'] = this_customer_items.iloc[idx, 1]
        ws[f'B{item_row_no}'] = this_customer_items.iloc[idx, 0]
        ws[f'C{item_row_no}'] = this_customer_items.iloc[idx, 3]
        ws[f'D{item_row_no}'] = this_customer_items.iloc[idx, 2]
        ws[f'E{item_row_no}'] = this_customer_items.iloc[idx, 4]
        ws[f'F{item_row_no}'] = this_customer_items.iloc[idx, 5]
        ws[f'G{item_row_no}'] = this_customer_items.iloc[idx, 6]
        ws[f'H{item_row_no}'] = this_customer_items.iloc[idx, 7]
        item_row_no += 1

    ws[f'A{item_row_no}'] = 'Total'
    ws[f'B{item_row_no}'] = f'=COUNTA(C6:C{item_row_no-1})'
    ws[f'F{item_row_no}'] = f'=SUM(F6:F{item_row_no-1})'
    ws[f'G{item_row_no}'] = f'=SUM(G6:G{item_row_no-1})'
    ws[f'H{item_row_no}'] = f'=SUM(H6:H{item_row_no-1})'

    ws = stylizeInvoiceForm(ws, 6, item_row_no)
    wb.save(f'./invoice/service_invoice_{customer}.xlsx')

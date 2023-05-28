from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border

# wb = Workbook()
# ws = wb.active # create ws object
# ws.title = 'Sheet1'
# ws.append(['name', 'age', 'gender'])
# ws.append(['Columbus', '55', 'male'])
# wb.save(filename='test.xlsx')
filepath = 'test.xlsx'
wb = load_workbook(filename=filepath)
# ws = wb.active
# ws.append(['Gozila', '40', 'male'])
# wb.create_sheet('Sheet2')
# wb.create_sheet('Sheet3')
# wb.save(filename=filepath)
# wb.copy_worksheet(wb['Sheet1'])
# wb.active = wb['Sheet1 Copy']
# wb = load_workbook('copy_sheet.xlsx')
# wb['Sheet1 Copy'].title = 'Sheet1 Clone'
# wb.move_sheet(wb['Sheet1 Clone'], -1)
# wb.save(filename='move_sheet.xlsx')
# wb = load_workbook('move_sheet.xlsx')
# wb.remove(wb['Sheet1 Clone'])
# wb.save(filename='remove_sheet.xlsx')
ws = wb['Sheet1']
# print(ws['A2'].value)
# print(ws[2][0].value) # row-2, col-0 cell
# new_data = ['Hellen', '45', 'female']
# row_no = 1
# for row in ws :
#     if row_no == 2 :
#         for cell in row :
#             cell.value = new_data[cell.col_idx-1]
#     row_no += 1
#         print(f'{cell} : {cell.value}', end = '\t')
#     print()
# ws['A1'] = 'first name'
# ws['A4'] = 'Hellen' # input data in cell
# ws['B4'] = '45'
# ws['C4'] = 'female'
#
# ws.append(['Solomon', '60', 'male']) # add data in next row
#
# new_data_list = [['Python', '16', 'female'], ['Kotlin', '60', 'male']]
# for new_data in new_data_list :
#     ws.append(new_data)
# ws.delete_rows(idx = 2, amount = 1) # delete 1 row from 3rd row
# ws.delete_cols(idx = 2, amount = 2) # delte 2 columns from 2nd(B) column

# src = ws['A2'].value
# ws['A5'] = src # copy value of A2 in A5
#
# for row in ws['A6:C6'] : # for roof range in A6 - C6
#     for cell in row :
#         cell.value = src

# src_data = []
# for cell in ws[2] :
#     src_data.append(cell.value)
#
# for row in ws['A5:C5'] :
#     for cell in row:
#         cell.value = src_data[cell.col_idx-1]

# ws.move_range('B1:C3', rows=1, cols=2)

# ws.merge_cells('A1:B1')
# ws.merge_cells('A2:B2')
#
# ws.unmerge_cells('A2:B2')

# ws['B2'].data_type = 'i'
# ws['B3'].data_type = 'i'
# ws['A4'] = 'age sum'
# ws['B4'] = '=sum(B2:B3)'

# fill_style = PatternFill(fill_type='solid', start_color='000000')
# font_style = Font(color='FFFFFF', sz=12, bold=True)
#
# for row in ws :
#     for cell in row :
#         if cell.row == 1 :
#             cell.fill = fill_style
#             cell.font = font_style

# for i in range(1, 6) :
#     ws.row_dimensions[i].height = 24
# ws.column_dimensions['A'].width = 14

# alignment_style = Alignment(horizontal='center', vertical='center')
# for row in ws :
#     for cell in row :
#         cell.alignment = alignment_style

side_style = Side(style='medium', color='000000')
border_styles = Border(
    left=side_style,
    right=side_style,
    top=side_style,
    bottom=side_style
)

for row in ws :
    for cell in row :
        cell.border = border_styles

wb.save(filename='test2.xlsx')

from openpyxl import load_workbook

wb = load_workbook(filename='./docs/customer_satisfaction.xlsx')
ws = wb['Sheet1']

for row in ws :
    for cell in row :
        if 0 < cell.col_idx < 7 and cell.row > 1 :
            cell.data_type = 'i'

this_max_row = ws.max_row
ws[f'A{this_max_row+1}'] = 'avg/sum'

for col_let in ['B', 'C', 'D'] :
    avg_formula = f'=AVERAGE({col_let}2:{col_let}{this_max_row})'
    ws[f'{col_let}{this_max_row + 1}'] = avg_formula
    ws[f'{col_let}{this_max_row + 1}'].number_format = '#.##'

for col_let in ['E', 'F'] :
    sum_formula = f'=SUM({col_let}2:{col_let}{this_max_row})'
    ws[f'{col_let}{this_max_row + 1}'] = sum_formula

wb.create_sheet('Resume')
ws = wb['Resume']

summary_table_form = [
    ['gender distribution', 'male', 'female'],
    ['customers', '', ''],
    ['rate', '', ''],
    ['', '', ''],
    ['age distribution', '20', '30', '40', 'over 50'],
    ['customers', '', ''],
    ['rate', '', ''],
    ['', '', ''],
    ['research result'],
    ['satisfaction'],
    ['willing to reorder'],
    ['intention to recommand'],
    ['visit count'],
    ['claim count']
]

for form_row in summary_table_form :
    ws.append(form_row)

# male/female count & rate
for gender in ['B', 'male'], ['C', 'female'] :
    ws[gender[0]+'2'] = f'=COUNTIF(Sheet1!H2:H{ws.max_row}, "{gender[1]}")'
    ws[gender[0]+'3'] = f'={gender[0]}2/SUM(B2:C2)'
    ws[gender[0]+'3'].number_format = '0.00%'

# per age count & rate
for age in [['B', '20'], ['C', '30'], ['D', 40], ['E', 'over 50']] :
    ws[age[0]+'6'] = f'=COUNTIF(Sheet1!G2:G{ws.max_row}, "{age[1]}")'
    ws[age[0]+'7'] = f'={age[0]}6/SUM(B6:E6)'
    ws[age[0]+'7'].number_format = '0.00%'

# average, sum
row_no = 10
for alphabet in ['B', 'C', 'D', 'E', 'F'] :
    ws['B'+str(row_no)] = f'=Sheet1!{alphabet}{ws.max_row+1}'
    if row_no < 13 :
        ws['B'+str(row_no)].number_format = '#.##'
    row_no += 1

ws.column_dimensions['A'].width = 15
for alphabet in ['B', 'C', 'D', 'E'] :
    ws.column_dimensions[alphabet].width = 10

for i in range(len(summary_table_form)+1) :
    ws.row_dimensions[i].height = 25

from openpyxl.styles import Alignment, Side, Border, Font, PatternFill, Color

alignment_style = Alignment(horizontal='center', vertical='center')
side = Side(color='CCCCCC', border_style='medium')
border_style = Border(
    left = side,
    right = side,
    top = side,
    bottom = side,
)

# set font & edge
style_apply_cells = ['A1:C3', 'A5:E7', 'A9:B14']
for cell_area in style_apply_cells :
    for row in ws[cell_area] :
        for cell in row :
            cell.alignment = alignment_style
            cell.border = border_style

header_font_style = Font(sz=12, bold=True)
header_background_style = PatternFill(fill_type='solid', start_color='ffd663')
style_apply_cells_for_header = ['A1:C1', 'A5:E5', 'A9:B9']

for cell_area in style_apply_cells_for_header :
    for row in ws[cell_area] :
        for cell in row :
            cell.font = header_font_style
            cell.fill = header_background_style

ws.merge_cells('A9:B9')

wb.save(filename='customer_satisfaction_report.xlsx')

from openpyxl import load_workbook

load_wb = load_workbook('students_list.xlsx')
load_ws = load_wb.active

name_list = []
birthday_list = []
code_list = []
for i in range(1, load_ws.max_row + 1) :
    name_list.append(load_ws.cell(i, 1).value)
    birthday_list.append(load_ws.cell(i, 2).value)
    code_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birthday_list)
print(code_list)